import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import os
import time

def scrape_data():
    print("Scraping Pikalytics data")
    options = webdriver.FirefoxOptions()
    options.headless = True
    driver = webdriver.Firefox(options=options)
    url = 'https://www.pikalytics.com/pokedex/gen9ou/'
    driver.get(url)

    # Allow Java to load
    time.sleep(3)

    # Locate scrollable element
    scroll_element = driver.find_element(By.CLASS_NAME, 'pokedex-wrapper-min')
    last_height = driver.execute_script("return arguments[0].scrollHeight", scroll_element)
    while True:
        driver.execute_script("arguments[0].scrollTo(0, arguments[0].scrollHeight);", scroll_element)
        time.sleep(0.10)
        new_height = driver.execute_script("return arguments[0].scrollHeight", scroll_element)

        if new_height == last_height:
            break

        last_height = new_height

    # Make request and parse HTML
    response = requests.get(url)
    print("Response status code:", response.status_code)
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, 'html.parser')

    # Close browser
    driver.quit()

    # Find pokemon on page
    pokemon_list = soup.find_all('a', class_='pokedex_entry')
    print("Found", len(pokemon_list), "pokemon.")

    data = []
    count = 0

    #Iterate over each item
    for pokemon in pokemon_list:
        
        try:
            pokemon_name = pokemon.find('span', class_='pokemon-name').text.strip()
        except AttributeError:
            pokemon_name = 'N/A'
        
        try:
            usage_percent_text = pokemon.find('span', class_='margin-right-20').text.strip()
            usage_percent = float(usage_percent_text.replace('%', ''))
        except (AttributeError, ValueError):
            usage_percent = 0.0
        
        if usage_percent >= 1.00:
            count += 1
            usage_percent = usage_percent / 100
            data.append([pokemon_name, usage_percent])

    print(count, 'entries found which matched criteria.')
    print('Scraping complete.')
    return data

def write_to_excel(data, output):
    print('Writing data to Excel...')
    # Delete previous if exists
    if os.path.exists(output):
        os.remove(output)
        print('Deleted previous file.')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write headers
    headers = ['Name', 'Usage']
    ws.append(headers)
    percent_style = NamedStyle(name="percent", number_format='0.00%')

    # Write data
    for row in data:
        ws.append(row)

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.style = percent_style


    # Adjust column width based on content
    for col in ws.columns:
        max_length = 0
        for cell in col:
            max_length = max(max_length, len(str(cell.value)))
        column_letter = get_column_letter(col[0].column)
        ws.column_dimensions[column_letter].width = max_length + 5.5

    # Save workbook
    wb.save(output)
    print("Data written successfully to:", output)

if __name__ == "__main__":
    output = 'pokemon_ou_usage.xlsx'

    print('Starting scraping process...')

    data = scrape_data()
    write_to_excel(data, output)

    print('Process completed successfully.')